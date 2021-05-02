from __future__ import print_function
from woocommerce import API
import json
import datetime
from datetime import datetime, timedelta
import pickle
import sqlite3
import os
from openpyxl import load_workbook
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request


class Customer:
    def __init__(self, first, last, phone, email,
                 street, city, state, zip_code, pickup_address, apt=None):
        self.first = first
        self.last = last
        self.phone = phone
        self.email = email
        self.street = street
        self.apt = apt
        self.city = city
        self.state = state
        self.zip_code = zip_code
        self.delivery_address = f'{street}, {city}, {state} {zip_code}'
        self.pickup_address = pickup_address

    def fullname(self):
        return f'{self.first} {self.last}'

    def __repr__(self):
        return f'Customer({self.first}, {self.last}, {self.phone}, {self.email},' \
               f' {self.street}, {self.city}, {self.state}, {self.zip_code}, {self.pickup_address})'

    def __str__(self):
        return f'''    {self.first} {self.last}
phone: {self.phone}
email: {self.email}
delivery: {self.delivery_address}
pickup: {self.pickup_address}
'''

    def save(self):
        # CONNECT TO THE DATABASE AND CREATE A SERVER
        conn = sqlite3.connect('customers.db')
        c = conn.cursor()

        c.execute(f"INSERT INTO customers VALUES ('{self.first}','{self.last}','{self.phone}','{self.email}')")
        conn.commit()
        conn.close()


class RentalOrder:
    total_orders = 0

    def __init__(self, invoice, date, customer, lg_boxes, xl_boxes, lg_dollies,
                 xl_dollies, labels, zip_ties, bins, rental_period,
                 delivery_date, delivery_time, pickup_date):
        self.invoice = invoice
        self.date = date
        self.customer = customer
        self.lg_boxes = int(lg_boxes)
        self.xl_boxes = int(xl_boxes)
        self.lg_dollies = int(lg_dollies)
        self.xl_dollies = int(xl_dollies)
        self.labels = int(labels)
        self.zip_ties = int(zip_ties)
        self.bins = int(bins)
        self.customer_signed = False
        self.employee_signed = False
        self.rental_period = rental_period
        self.delivery_date = delivery_date
        self.delivery_time = delivery_time
        self.pickup_date = pickup_date

    def __repr__(self):
        return f'RentalOrder({self.invoice}, {self.date}, {self.customer}, {self.lg_boxes}, {self.xl_boxes},' \
               f'{self.lg_dollies},{self.xl_dollies}, {self.labels}, {self.zip_ties},' \
               f' {self.bins}, {self.customer_signed}. {self.employee_signed}, {self.rental_period},' \
               f' {self.delivery_date}, {self.delivery_time}, {self.pickup_date})'

    def __str__(self):
        return f''' 
Invoice: {self.invoice}
Date Ordered: {self.date}
Lg Boxes: {self.lg_boxes}
Xl Boxes: {self.xl_boxes}
Lg Dollies: {self.lg_dollies}
Xl Dollies: {self.xl_dollies}
Labels: {self.labels}
Zip-Ties: {self.zip_ties}
Bins: {self.bins}
Customer Signed: {self.customer_signed}
Employee Signed: {self.employee_signed}
Rental Period: {self.rental_period}
Delivery Date: {self.delivery_date}
Pick Up Date: {self.pickup_date}
'''

    def save(self):
        # CONNECT TO THE DATABASE AND CREATE A SERVER
        conn = sqlite3.connect('rental_orders.db')
        c = conn.cursor()

        c.execute(f'''INSERT INTO rental_orders VALUES 
('{self.invoice}','{self.customer}',
'{self.lg_boxes}','{self.xl_boxes}',
'{self.lg_dollies}', '{self.xl_dollies}',
'{self.labels}', '{self.zip_ties}',
'{self.bins}', '{self.customer_signed}',
'{self.employee_signed}','{self.rental_period}',
'{self.delivery_date}', '{self.pickup_date}')
''')
        conn.commit()
        conn.close()


def get_customer():

    current_order = last_order
    f_name = current_order['billing']['first_name']
    l_name = current_order['billing']['last_name']
    phone = current_order['billing']['phone']
    email = current_order['billing']['email']
    street = current_order['shipping']['address_1']
    apt = current_order['shipping']['address_2']
    city = current_order['shipping']['city']
    state = current_order['shipping']['state']
    zip_code = current_order['shipping']['postcode']
    pickup_address = current_order['meta_data'][4]['value']

    customer = Customer(first=f_name, last=l_name, phone=phone, email=email,
                        street=street, apt=apt, city=city, state=state,
                        zip_code=zip_code, pickup_address=pickup_address)

    # current_customer.save()
    print(json.dumps(current_order, indent=4))
    return customer


def get_order():
    customer = get_customer()
    current_order = last_order
    invoice = current_order['id']
    date = datetime.strptime(current_order['date_created'], '%Y-%m-%dT%H:%M:%S').date()

    delivery_time = current_order['meta_data'][0]['value'] + current_order['meta_data'][1]['value']

    try:
        delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%Y-%m-%d').date()
    except ValueError:
        delivery_date = datetime.strptime(current_order['meta_data'][0]['value'], '%m/%d/%Y').date()

    try:
        delivery_time = datetime.strptime(delivery_time, '%Y-%m-%d%H:%M')
    except ValueError:
        delivery_time = datetime.strptime(delivery_time, '%m/%d/%Y%H:%M %p')

    # THIS LOCATES THE RENTAL PERIOD AND CREATES PICKUP DATETIME OBJECT
    try:
        rental_period = int(current_order['line_items'][0]['meta_data'][0]['value'][0])
        if rental_period == 1:
            pickup_date = delivery_date + timedelta(days=7)
            rental_period = '1 Week'
        elif rental_period == 2:
            pickup_date = delivery_date + timedelta(days=14)
            rental_period = '2 Weeks'
        elif rental_period == 3:
            pickup_date = delivery_date + timedelta(days=21)
            rental_period = '3 Weeks'
        elif rental_period == 4:
            pickup_date = delivery_date + timedelta(days=28)
            rental_period = '4 Weeks'
    except IndexError:
        pickup_date = delivery_date + timedelta(days=7)
        rental_period = '1 Week'

    # THIS LOCATES WOO-COMMERCE PRODUCT ID AND CREATES BOX PACKAGE ACCORDINGLY
    # SETS VALUES FOR LG BOXES, XL BOXES, ETC
    for x in range(len(current_order['line_items'])):
        if current_order['line_items'][x]['product_id'] == 1270:
            lg_boxes = 70
            xl_boxes = 10
            lg_dollies = 4
            xl_dollies = 2
            labels = 80
            zip_ties = 80
            bins = 0
            break
        elif current_order['line_items'][x]['product_id'] == 1515:
            lg_boxes = 50
            xl_boxes = 10
            lg_dollies = 3
            xl_dollies = 1
            labels = 60
            zip_ties = 60
            bins = 0
            break
        elif current_order['line_items'][x]['product_id'] == 1510:
            lg_boxes = 35
            xl_boxes = 5
            lg_dollies = 2
            xl_dollies = 0
            labels = 40
            zip_ties = 40
            bins = 0
        elif current_order['line_items'][x]['product_id'] == 1505:
            lg_boxes = 18
            xl_boxes = 2
            lg_dollies = 1
            xl_dollies = 0
            labels = 20
            zip_ties = 20
            bins = 0
            break
        elif current_order['line_items'][x]['product_id'] == 1545:
            lg_boxes = 1
            xl_boxes = 0
            lg_dollies = 0
            xl_dollies = 0
            labels = 0
            zip_ties = 0
            bins = 0
            break
        elif current_order['line_items'][x]['product_id'] == 1291:
            lg_boxes = 0
            xl_boxes = 0
            lg_dollies = 0
            xl_dollies = 0
            labels = 0
            zip_ties = 0
            bins = current_order['line_items'][x]['quantity']
            break
        else:
            lg_boxes = 0
            xl_boxes = 0
            lg_dollies = 0
            xl_dollies = 0
            labels = 0
            zip_ties = 0
            bins = 0

    delivery_time = delivery_time.strftime('%Y-%m-%dT%H:%M:%S-04:00')
    pickup_time = pickup_date.strftime('%Y-%m-%dT%H:%M:%S-04:00')

    order = RentalOrder(invoice, date=date, customer=customer.fullname, lg_boxes=lg_boxes, xl_boxes=xl_boxes,
                        lg_dollies=lg_dollies, xl_dollies=xl_dollies, labels=labels, zip_ties=zip_ties, bins=bins,
                        rental_period=rental_period, delivery_date=delivery_date,
                        delivery_time=delivery_time, pickup_date=pickup_date)

    # c_order.save()

    return order


def get_events():
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('calendar', 'v3', credentials=creds)

    cal = service.calendars().get(calendarId='primary').execute()
    page_token = None
    while True:
        events = service.events().list(calendarId='primary', pageToken=page_token).execute()
        for e in events['items']:
            print(e['summary'])
        page_token = events.get('nextPageToken')
        if not page_token:
            break


def post_events(customer, order):

    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('calendar', 'v3', credentials=creds)

    end_time = datetime.strptime(order.delivery_time, '%Y-%m-%dT%H:%M:%S-04:00') + timedelta(hours=1)
    end_time = end_time.strftime('%Y-%m-%dT%H:%M:%S-04:00')

    event = {
        'summary': f'{customer.first} {customer.last} Box Delivery',
        'location': f'{customer.delivery_address}',
        'description': f'''{customer.fullname}
phone: {customer.phone}
address: {customer.delivery_address}
email: {customer.email}
{order.lg_boxes} Lg Boxes
{order.xl_boxes} Xl Boxes
{order.lg_dollies} Dollies
{order.xl_dollies} Xl Dollies
{order.labels} Labels & Zip Ties
{order.bins} Bins''',

        'start': {
            'dateTime': f'{order.delivery_time}',
            'timeZone': 'America/New_York',
        },
        'end': {
            'dateTime': f'{end_time}',
            'timeZone': 'America/New_York',
        },
        'reminders': {
            'useDefault': False,
            'overrides': [
                {'method': 'email', 'minutes': 24 * 60},
                {'method': 'popup', 'minutes': 10},
            ],
        },
    }

    # Call the Calendar API
    event = service.events().insert(calendarId='primary', body=event).execute()

    pickup_time = order.pickup_date.strftime('%Y-%m-%dT%H:%M:%S-04:00')

    end_time = datetime.strptime(pickup_time, '%Y-%m-%dT%H:%M:%S-04:00') + timedelta(hours=1)
    end_time = end_time.strftime('%Y-%m-%dT%H:%M:%S-04:00')

    pickup_event = {
        'summary': f'{customer.first} {customer.last} Box Pick Up',
        'location': f'{customer.pickup_address}',
        'description': f'''{customer.fullname}
phone: {customer.phone}
Address {customer.pickup_address}
email: {customer.email}
{order.lg_boxes} Lg Boxes                                                                      
{order.xl_boxes} Xl Boxes
{order.lg_dollies} Dollies
{order.xl_dollies} Xl Dollies
{order.labels} Labels & Zip Ties
{order.bins} Bins''',

        'start': {
            'dateTime': f'{pickup_time}',
            'timeZone': 'America/New_York',
        },
        'end': {
            'dateTime': f'{end_time}',
            'timeZone': 'America/New_York',
        },
        'reminders': {
            'useDefault': False,
            'overrides': [
                {'method': 'email', 'minutes': 24 * 60},
                {'method': 'popup', 'minutes': 10},
            ],
        },
    }
    pickup_event = service.events().insert(calendarId='primary', body=pickup_event).execute()
    print('Delivery Event created: %s' % (event.get('htmlLink')))
    print('Pick-up Event created: %s' % (pickup_event.get('htmlLink')))


def save_order(obj, filename):
    with open(filename, 'wb') as output:  # Overwrites any existing file.
        pickle.dump(obj, output, pickle.HIGHEST_PROTOCOL)


SCOPES = ['https://www.googleapis.com/auth/calendar']

# WOO-COMMERCE API CREDENTIALS
wcapi = API(
    url=os.environ.get('WCAPI_URL'),
    consumer_key=os.environ.get('WCAPI_CONSUMER_KEY'),
    consumer_secret=os.environ.get('WCAPI_CONSUMER_SECRET'),
    version="wc/v3"
)

# COLLECT ORDER DATA
orders = wcapi.get('orders')
data = orders.json()
last_order = data[0]
new_order = True


def write_workbook(order, customer):
    filename = f"{customer.first} {customer.last}{order.invoice}.xlsx"

    path = os.path.dirname(__file__)
    file = os.path.join(path, 'rental_order.xlsx')
    wb = load_workbook(file)
    ws = wb.active

    ws['G3'] = order.date
    ws['G4'] = order.invoice
    ws['C7'] = f'{customer.first} {customer.last}'
    ws['C8'] = f'{customer.street} {customer.apt}'
    ws['C9'] = f'{customer.city}, {customer.state} {customer.zip_code}'
    ws['C10'] = customer.phone
    ws['C11'] = customer.email
    ws['F7'] = f'{customer.street} {customer.apt}'
    ws['F8'] = f'{customer.city}, {customer.state} {customer.zip_code}'
    ws['F11'] = customer.pickup_address
    ws['F16'] = order.delivery_date
    ws['G16'] = order.pickup_date
    ws['B18'] = order.rental_period
    # ws['G18'] = order.was_delivered()
    ws['B21'] = order.lg_boxes
    ws['B22'] = order.xl_boxes
    ws['B23'] = order.lg_dollies
    ws['B24'] = order.xl_dollies
    ws['B25'] = order.labels
    ws['B26'] = order.zip_ties
    ws['B27'] = order.bins

    wb.save(filename)
    return wb


def main():

    customer = get_customer()
    order = get_order()
    print(customer)
    print(order)
    write_workbook(order, customer)
    # post = input(f'Would you like to post delivery and pick-up events for {customer.first} {customer.last}?\n'
    #              f'Please Enter y/n: ')
    # if 'y' in post.lower():
    #     post_events(customer, order)
    post_events(customer, order)


if __name__ == '__main__':
    main()

