# COLLECT EMAIL LIST #

from woocommerce import API
import json
from openpyxl import Workbook, load_workbook
import os

#1. First I need to connect to woocommerce API

wcapi = API(
    url=os.environ.get('WCAPI_URL'),
    consumer_key=os.environ.get('WCAPI_CONSUMER_KEY'),
    consumer_secret=os.environ.get('WCAPI_CONSUMER_SECRET'),
    version="wc/v3"
    )

#2. Then I collect the order data from the woocommerce API

def list_emails():
    
    count = 0
    pages = 25
    emails_dict = {}

    for num in range(1, pages):
        orders = wcapi.get('orders', params={'per_page': 100, 'page':num}).json()
        for i in orders:
            count += 1
            name = i['billing']['first_name'] + i['billing']['last_name']
            email = i['billing']['email']
            emails_dict[name] = email
         #   print(f'{name}: {email}')
        if len(orders) <= 0:
            return emails_dict

    return emails_dict      

# CREATE EXCEL WORKBOOK
def create_workbook():
    # emails = list_emails()
    wb = load_workbook('customer-email-list.xlsx')
    ws = wb.active
    print(ws)

    ws['A2'].value = 'Mick Jagger'

    filename = 'customer-email-list.xlsx'

    # key_list = emails.keys()
    # for row in range(1, len(key_list)):
    #     ws.append(emails[row])
    
    wb.save(filename)

create_workbook()